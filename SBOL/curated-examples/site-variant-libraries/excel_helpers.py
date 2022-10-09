from pathlib import Path

import openpyxl
from openpyxl.worksheet import cell_range, worksheet

VARIANTS_SHEET = 'Amino Acid Variants'
LIBRARY_NAME_CELL = 'C10'
FIRST_AMINO_ACID_COLUMN = 'F'
ORIGINAL_AMINO_ACID_ROW = 12
FIRST_VARIANT_ROW = 14
LAST_VARIANT_ROW = 35


def get_column_number(column_letter: str) -> int:
    """Turn a column letter (e.g., C) into a number (e.g., 3)

    :param column_letter: column to convert
    :return: integer equivalent of column_letter
    """
    return cell_range.range_boundaries(f'{column_letter}1')[0]


def row_ends(sheet: worksheet, row: int, min_col: int = 1) -> int:
    """Find the column at which a row ends

    :param sheet: Sheet to search
    :param row: Row to search
    :param min_col: Column on which to start, defaults to first column
    :return: numerical value of last column containing data
    """
    # convert min column to numerical to determine starting point
    column_iterator = sheet.iter_cols(min_row=row, max_row=row, min_col=min_col, values_only=True)
    max_sequence_length = 100000  # TODO: use a proper generator counter rather than a big range
    return next(i-1 for v, i in zip(column_iterator, range(min_col, max_sequence_length)) if not v[0])


def read_variant_table(excel_file: Path) -> tuple[str, str, list[list]]:
    """Extract an amino acid site-variant table from a Twist amino-acid site-variant library Excel sheet

    :param excel_file: location of file to read
    :return: Tuple of library name, original sequence, list of variant lists for each site, in order
    """
    print(f'Loading workbook "{excel_file}"')
    work_book = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = work_book[VARIANTS_SHEET]

    # First, get the library name
    library_name = sheet[LIBRARY_NAME_CELL].value
    print(f'Library is named "{library_name}"')

    # Then get the base sequence
    print('Extracting base sequence')
    first_aa_column = get_column_number(FIRST_AMINO_ACID_COLUMN)
    last_aa_column = row_ends(sheet, ORIGINAL_AMINO_ACID_ROW, first_aa_column)
    # Get row from sheet and concatenate it into a string
    row_iterator = sheet.iter_rows(min_row=ORIGINAL_AMINO_ACID_ROW, max_row=ORIGINAL_AMINO_ACID_ROW,
                                   min_col=first_aa_column, max_col=last_aa_column, values_only=True)
    base_sequence = [''.join(row) for row in row_iterator][0]
    print(f'Found sequence {len(base_sequence)} residues long: "{base_sequence}"')

    # Finally, get all the variant lists
    column_iterator = sheet.iter_cols(min_row=FIRST_VARIANT_ROW, max_row=LAST_VARIANT_ROW,
                                      min_col=first_aa_column, max_col=last_aa_column, values_only=True)
    variant_lists = [[v for v in column if v] for column in column_iterator]  # drop the empty cells from each range
    print(f'Library specifies a total of {sum(len(variants) for variants in variant_lists)} variants')

    return library_name, base_sequence, variant_lists
